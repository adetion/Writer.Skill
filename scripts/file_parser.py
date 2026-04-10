#!/usr/bin/env python3
import os
import yaml
import re
from typing import List, Dict, Any

class FileParser:
    SUPPORTED_FORMATS = [".docx", ".xlsx", ".pdf", ".txt", ".md"]
    
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def _get_file_type(self, file_path: str) -> str:
        """获取文件类型"""
        ext = os.path.splitext(file_path)[1].lower()
        if ext in self.SUPPORTED_FORMATS:
            return ext.lstrip(".")
        raise ValueError(f"不支持的文件格式: {ext}，支持格式: {','.join(self.SUPPORTED_FORMATS)}")
    
    def parse_txt(self, file_path: str) -> List[Dict[str, Any]]:
        """解析txt/md文件"""
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 按标题拆分段落
        entries = []
        sections = re.split(r'\n#{1,6}\s+', content)
        if not sections[0].strip():
            sections = sections[1:]
        
        for i, section in enumerate(sections):
            lines = section.split("\n", 1)
            title = lines[0].strip() if len(lines) > 0 else f"条目_{i+1}"
            content = lines[1].strip() if len(lines) > 1 else section.strip()
            
            if not content:
                continue
            
            # 自动提取标签
            tags = []
            if "人物" in title or "角色" in title:
                tags.append("人物")
            if "世界观" in title or "设定" in title:
                tags.append("世界观")
            if "剧情" in title or "大纲" in title:
                tags.append("剧情")
            
            entries.append({
                "id": f"auto_{os.path.basename(file_path).split('.')[0]}_{i+1}",
                "title": title,
                "content": content,
                "tags": tags,
                "priority": 8 if tags else 5
            })
        return entries
    
    def parse_docx(self, file_path: str) -> List[Dict[str, Any]]:
        """解析docx文件"""
        try:
            from docx import Document
        except ImportError:
            raise RuntimeError("缺少docx解析依赖，请安装: pip install python-docx")
        
        doc = Document(file_path)
        content = "\n".join([para.text for para in doc.paragraphs if para.text.strip()])
        return self.parse_txt_content(content, os.path.basename(file_path))
    
    def parse_pdf(self, file_path: str) -> List[Dict[str, Any]]:
        """解析pdf文件"""
        try:
            import PyPDF2
        except ImportError:
            raise RuntimeError("缺少pdf解析依赖，请安装: pip install pypdf2")
        
        content = ""
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                content += page.extract_text() + "\n"
        return self.parse_txt_content(content, os.path.basename(file_path))
    
    def parse_xlsx(self, file_path: str) -> List[Dict[str, Any]]:
        """解析xlsx文件"""
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("缺少excel解析依赖，请安装: pip install openpyxl")
        
        wb = openpyxl.load_workbook(file_path)
        entries = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append([str(cell) if cell is not None else "" for cell in row])
            
            if not data:
                continue
            
            # 第一行为表头，后续为内容
            headers = data[0]
            for i, row in enumerate(data[1:], start=1):
                row_dict = dict(zip(headers, row))
                content = "\n".join([f"{k}: {v}" for k, v in row_dict.items() if v.strip()])
                title = row_dict.get("名称", row_dict.get("title", f"{sheet_name}_条目_{i}"))
                
                tags = [sheet_name]
                if "人物" in sheet_name:
                    tags.append("人物")
                if "设定" in sheet_name:
                    tags.append("世界观")
                
                entries.append({
                    "id": f"auto_xlsx_{sheet_name}_{i}",
                    "title": title,
                    "content": content,
                    "tags": tags,
                    "priority": 8
                })
        return entries
    
    def parse_txt_content(self, content: str, filename: str) -> List[Dict[str, Any]]:
        """通用文本内容解析"""
        entries = []
        # 按空行分段
        paragraphs = [p.strip() for p in re.split(r'\n\s*\n', content) if p.strip()]
        
        for i, para in enumerate(paragraphs):
            # 提取第一句作为标题
            lines = para.split("。", 1) if "。" in para else para.split("\n", 1)
            title = lines[0].strip()[:50] if len(lines) > 0 else f"{filename.split('.')[0]}_条目_{i+1}"
            content = para.strip()
            
            tags = []
            if "人物" in title or "姓名" in content:
                tags.append("人物")
            if "境界" in content or "世界" in content or "规则" in content:
                tags.append("世界观")
            if "剧情" in title or "情节" in content:
                tags.append("剧情")
            
            entries.append({
                "id": f"auto_{filename.split('.')[0]}_{i+1}",
                "title": title,
                "content": content,
                "tags": tags,
                "priority": 7 if tags else 5
            })
        return entries
    
    def parse_file(self, file_path: str, output_yaml: str = None) -> str:
        """解析文件并保存为YAML格式知识库"""
        file_type = self._get_file_type(file_path)
        parse_func = getattr(self, f"parse_{file_type}")
        entries = parse_func(file_path)
        
        if not output_yaml:
            output_yaml = os.path.join(self.output_dir, f"{os.path.splitext(os.path.basename(file_path))[0]}.yaml")
        
        with open(output_yaml, 'w', encoding='utf-8') as f:
            yaml.dump(entries, f, allow_unicode=True, sort_keys=False)
        
        return output_yaml, len(entries)

if __name__ == "__main__":
    import sys
    if len(sys.argv) != 2:
        print("用法: python file_parser.py <待解析文件路径>")
        sys.exit(1)
    
    parser = FileParser(output_dir="./")
    output_path, count = parser.parse_file(sys.argv[1])
    print(f"✅ 解析完成，共生成{count}条知识库条目，保存到: {output_path}")
